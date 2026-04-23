"""
Usage:
    python run_reports.py [--host HOST] [--port PORT] [--user USER]
                         [--password PASSWORD] [--database DATABASE]
    Or set environment variables:
        DB_HOST, DB_PORT, DB_USER, DB_PASSWORD, DB_NAME

Requirements:
    pip install mysql-connector-python pandas openpyxl jinja2 python-dotenv
"""

import os
import sys
import logging
import argparse
import textwrap
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import pandas as pd
import mysql.connector
from mysql.connector import Error as MySQLError

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Output directory ─────────────────────────────────────────────────────────
REPORTS_DIR = Path("reports")
REPORTS_DIR.mkdir(exist_ok=True)

# ═════════════════════════════════════════════════════════════════════════════
#  QUERY DEFINITIONS
#  Each entry: (short_id, human_label, sql_string)
# ═════════════════════════════════════════════════════════════════════════════
QUERIES = [

    # ── Q1 ───────────────────────────────────────────────────────────────────
    ("Q01_total_customers",
     "Total Number of Customers",
     """
     SELECT COUNT(DISTINCT customer_identifier) AS total_customers
     FROM customer_detail
     """),

    ("Q02_gender_distribution",
     "Gender Distribution",
     """
     SELECT
         customer_gender,
         COUNT(*) AS customer_count,
         ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER (), 2) AS percentage
     FROM customer_detail
     GROUP BY customer_gender
     ORDER BY customer_count DESC
     """),

    ("Q03_customers_by_country",
     "Customer Count by Country",
     """
     SELECT
         customer_country,
         COUNT(*) AS total_customers
     FROM customer_detail
     GROUP BY customer_country
     ORDER BY total_customers DESC
     """),

    ("Q04_customers_by_status",
     "Customer Count by Status",
     """
     SELECT
         customer_status,
         COUNT(*) AS count
     FROM customer_detail
     GROUP BY customer_status
     ORDER BY count DESC
     """),

    ("Q05_customers_by_type",
     "Customer Count by Type",
     """
     SELECT
         customer_type,
         COUNT(*) AS count
     FROM customer_detail
     GROUP BY customer_type
     ORDER BY count DESC
     """),

    ("Q06_customers_by_language",
     "Customer Count by Language Preference",
     """
     SELECT
         customer_lang,
         COUNT(*) AS count
     FROM customer_detail
     GROUP BY customer_lang
     ORDER BY count DESC
     """),

    ("Q07_gender_by_country",
     "Gender Distribution by Country",
     """
     SELECT
         customer_country,
         customer_gender,
         COUNT(*) AS count
     FROM customer_detail
     GROUP BY customer_country, customer_gender
     ORDER BY customer_country, count DESC
     """),

    ("Q08_type_by_status",
     "Customer Count by Type and Status",
     """
     SELECT
         customer_type,
         customer_status,
         COUNT(*) AS count
     FROM customer_detail
     GROUP BY customer_type, customer_status
     ORDER BY customer_type, count DESC
     """),

    ("Q09_new_customers_per_year",
     "New Customers Registered Each Year",
     """
     SELECT
         YEAR(audit_acceptance_time)  AS year,
         COUNT(*)                     AS new_customers
     FROM customer_detail
     GROUP BY YEAR(audit_acceptance_time)
     ORDER BY year DESC
     """),

    ("Q10_monthly_new_customers_3yr",
     "Monthly New Customers (Last 3 Years)",
     """
     SELECT
         YEAR(audit_acceptance_time)  AS year,
         MONTH(audit_acceptance_time) AS month,
         COUNT(*)                     AS new_customers
     FROM customer_detail
     WHERE audit_acceptance_time >= DATE_SUB(NOW(), INTERVAL 3 YEAR)
     GROUP BY YEAR(audit_acceptance_time), MONTH(audit_acceptance_time)
     ORDER BY year DESC, month DESC
     """),

    ("Q11_age_distribution",
     "Customer Age Distribution (Brackets)",
     """
     SELECT
         CASE
             WHEN TIMESTAMPDIFF(YEAR, customer_dob, CURDATE()) < 18       THEN 'Under 18'
             WHEN TIMESTAMPDIFF(YEAR, customer_dob, CURDATE()) BETWEEN 18 AND 25 THEN '18-25'
             WHEN TIMESTAMPDIFF(YEAR, customer_dob, CURDATE()) BETWEEN 26 AND 35 THEN '26-35'
             WHEN TIMESTAMPDIFF(YEAR, customer_dob, CURDATE()) BETWEEN 36 AND 45 THEN '36-45'
             WHEN TIMESTAMPDIFF(YEAR, customer_dob, CURDATE()) BETWEEN 46 AND 60 THEN '46-60'
             ELSE '60+'
         END AS age_group,
         COUNT(*) AS count
     FROM customer_detail
     WHERE customer_dob IS NOT NULL
     GROUP BY age_group
     ORDER BY age_group
     """),

    ("Q12_avg_age_by_country",
     "Average Customer Age by Country",
     """
     SELECT
         customer_country,
         ROUND(AVG(TIMESTAMPDIFF(YEAR, customer_dob, CURDATE())), 1) AS avg_age
     FROM customer_detail
     WHERE customer_dob IS NOT NULL
     GROUP BY customer_country
     ORDER BY avg_age DESC
     """),

    ("Q13_avg_age_by_gender",
     "Average Customer Age by Gender",
     """
     SELECT
         customer_gender,
         ROUND(AVG(TIMESTAMPDIFF(YEAR, customer_dob, CURDATE())), 1) AS avg_age
     FROM customer_detail
     WHERE customer_dob IS NOT NULL
     GROUP BY customer_gender
     """),

    ("Q14_customers_3yr_vs_before",
     "Customers Added: Last 3 Years vs Before",
     """
     SELECT
         CASE
             WHEN audit_acceptance_time >= DATE_SUB(NOW(), INTERVAL 3 YEAR)
                  THEN 'Last 3 Years'
             ELSE 'Before Last 3 Years'
         END AS period,
         COUNT(*) AS count
     FROM customer_detail
     GROUP BY period
     """),

    ("Q15_identification_type_dist",
     "Customer Identification Type Distribution",
     """
     SELECT
         customer_identification_type,
         COUNT(*) AS count
     FROM customer_identification
     GROUP BY customer_identification_type
     ORDER BY count DESC
     """),

    ("Q16_multiple_id_types",
     "Customers with Multiple Identification Types",
     """
     SELECT
         customer_identifier,
         COUNT(*) AS id_count
     FROM customer_identification
     GROUP BY customer_identifier
     HAVING id_count > 1
     ORDER BY id_count DESC
     """),

    ("Q17_proof_of_id_type_dist",
     "Proof of ID Document Type Distribution",
     """
     SELECT
         customer_proof_of_id_type,
         COUNT(*) AS count
     FROM customer_proof_of_id
     GROUP BY customer_proof_of_id_type
     ORDER BY count DESC
     """),

    ("Q18_active_proof_of_id",
     "Currently Active Proof of ID by Type",
     """
     SELECT
         customer_proof_of_id_type,
         COUNT(*) AS active_count
     FROM customer_proof_of_id
     WHERE start_date <= CURDATE()
       AND (end_date IS NULL OR end_date >= CURDATE())
     GROUP BY customer_proof_of_id_type
     ORDER BY active_count DESC
     """),

    ("Q19_expired_proof_of_id",
     "Customers with Expired Proof of ID",
     """
     SELECT COUNT(DISTINCT customer_identifier) AS customers_with_expired_id
     FROM customer_proof_of_id
     WHERE end_date < CURDATE()
     """),

    ("Q20_contact_type_dist",
     "Contact Type Distribution",
     """
     SELECT
         customer_contact_type,
         COUNT(*) AS count
     FROM customer_contact_information
     GROUP BY customer_contact_type
     ORDER BY count DESC
     """),

    ("Q21_active_contact_info",
     "Active Contact Information by Type",
     """
     SELECT
         customer_contact_type,
         COUNT(*) AS active_contacts
     FROM customer_contact_information
     WHERE start_date <= CURDATE()
       AND (end_date IS NULL OR end_date >= CURDATE())
     GROUP BY customer_contact_type
     """),

    ("Q22_multiple_contact_types",
     "Customers with Multiple Contact Types",
     """
     SELECT
         customer_identifier,
         COUNT(DISTINCT customer_contact_type) AS contact_type_count
     FROM customer_contact_information
     GROUP BY customer_identifier
     HAVING contact_type_count > 1
     ORDER BY contact_type_count DESC
     """),

    ("Q23_address_type_dist",
     "Address Type Distribution",
     """
     SELECT
         customer_address_type,
         COUNT(*) AS count
     FROM customer_address
     GROUP BY customer_address_type
     ORDER BY count DESC
     """),

    ("Q24_multiple_address_types",
     "Customers with Multiple Address Types",
     """
     SELECT
         customer_identifier,
         COUNT(DISTINCT customer_address_type) AS address_type_count
     FROM customer_address
     GROUP BY customer_identifier
     HAVING address_type_count > 1
     ORDER BY address_type_count DESC
     """),

    ("Q25_classification_type_dist",
     "Classification Type Distribution",
     """
     SELECT
         customer_classification_type,
         COUNT(*) AS count
     FROM customer_classification_type
     GROUP BY customer_classification_type
     ORDER BY count DESC
     """),

    ("Q26_classification_value_breakdown",
     "Classification Value Breakdown per Type",
     """
     SELECT
         customer_classification_type,
         customer_classification_value,
         COUNT(*) AS count
     FROM customer_classification_type
     GROUP BY customer_classification_type, customer_classification_value
     ORDER BY customer_classification_type, count DESC
     """),

    ("Q27_name_type_dist",
     "Customer Name Type Distribution",
     """
     SELECT
         customer_name_type,
         COUNT(*) AS count
     FROM customer_name
     GROUP BY customer_name_type
     ORDER BY count DESC
     """),

    ("Q28_multiple_name_types",
     "Customers with Multiple Name Types",
     """
     SELECT
         customer_identifier,
         COUNT(DISTINCT customer_name_type) AS name_type_count
     FROM customer_name
     GROUP BY customer_identifier
     HAVING name_type_count > 1
     ORDER BY name_type_count DESC
     """),

    ("Q29_gender_by_customer_type",
     "Gender Ratio per Customer Type",
     """
     SELECT
         customer_type,
         customer_gender,
         COUNT(*) AS count
     FROM customer_detail
     GROUP BY customer_type, customer_gender
     ORDER BY customer_type, count DESC
     """),

    ("Q30_country_gender_pct",
     "Country-wise Gender Breakdown with Percentage",
     """
     SELECT
         customer_country,
         customer_gender,
         COUNT(*) AS count,
         ROUND(
             COUNT(*) * 100.0
             / SUM(COUNT(*)) OVER (PARTITION BY customer_country),
             2
         ) AS pct_within_country
     FROM customer_detail
     GROUP BY customer_country, customer_gender
     ORDER BY customer_country, count DESC
     """),

    ("Q31_yoy_growth_rate",
     "Year-over-Year Customer Growth Rate (Last 3 Years)",
     """
     SELECT
         YEAR(audit_acceptance_time)                                         AS year,
         COUNT(*)                                                             AS new_customers,
         LAG(COUNT(*)) OVER (ORDER BY YEAR(audit_acceptance_time))           AS prev_year_count,
         ROUND(
             (COUNT(*) - LAG(COUNT(*)) OVER (ORDER BY YEAR(audit_acceptance_time)))
             * 100.0
             / NULLIF(LAG(COUNT(*)) OVER (ORDER BY YEAR(audit_acceptance_time)), 0),
             2
         )                                                                    AS growth_pct
     FROM customer_detail
     WHERE audit_acceptance_time >= DATE_SUB(NOW(), INTERVAL 3 YEAR)
     GROUP BY YEAR(audit_acceptance_time)
     ORDER BY year
     """),

    ("Q32_quarterly_registrations",
     "New Customers per Quarter (Last 3 Years)",
     """
     SELECT
         YEAR(audit_acceptance_time)    AS year,
         QUARTER(audit_acceptance_time) AS quarter,
         COUNT(*)                       AS new_customers
     FROM customer_detail
     WHERE audit_acceptance_time >= DATE_SUB(NOW(), INTERVAL 3 YEAR)
     GROUP BY YEAR(audit_acceptance_time), QUARTER(audit_acceptance_time)
     ORDER BY year, quarter
     """),

    ("Q33_top10_countries",
     "Top 10 Countries by Customer Count",
     """
     SELECT
         customer_country,
         COUNT(*) AS total
     FROM customer_detail
     GROUP BY customer_country
     ORDER BY total DESC
     LIMIT 10
     """),

    ("Q34_status_by_country",
     "Customer Status by Country",
     """
     SELECT
         customer_country,
         customer_status,
         COUNT(*) AS count
     FROM customer_detail
     GROUP BY customer_country, customer_status
     ORDER BY customer_country, count DESC
     """),

    ("Q35_language_pct_by_country",
     "Language Preference Percentage per Country",
     """
     SELECT
         customer_country,
         customer_lang,
         COUNT(*) AS count,
         ROUND(
             COUNT(*) * 100.0
             / SUM(COUNT(*)) OVER (PARTITION BY customer_country),
             2
         ) AS pct_within_country
     FROM customer_detail
     GROUP BY customer_country, customer_lang
     ORDER BY customer_country, count DESC
     """),

    ("Q36_active_vs_inactive_by_country",
     "Active vs Inactive Customers per Country",
     """
     SELECT
         customer_country,
         SUM(CASE WHEN customer_status = 'ACTIVE' THEN 1 ELSE 0 END) AS active,
         SUM(CASE WHEN customer_status != 'ACTIVE' THEN 1 ELSE 0 END) AS inactive,
         COUNT(*) AS total
     FROM customer_detail
     GROUP BY customer_country
     ORDER BY total DESC
     """),

    ("Q37_customers_no_contact",
     "Customers with No Contact Information",
     """
     SELECT COUNT(*) AS customers_without_contact
     FROM customer_detail cd
     WHERE NOT EXISTS (
         SELECT 1
         FROM customer_contact_information cci
         WHERE cci.customer_identifier = cd.customer_identifier
     )
     """),

    ("Q38_customers_no_address",
     "Customers with No Address on Record",
     """
     SELECT COUNT(*) AS customers_without_address
     FROM customer_detail cd
     WHERE NOT EXISTS (
         SELECT 1
         FROM customer_address ca
         WHERE ca.customer_identifier = cd.customer_identifier
     )
     """),

    ("Q39_customers_no_identification",
     "Customers with No Identification Document",
     """
     SELECT COUNT(*) AS customers_without_identification
     FROM customer_detail cd
     WHERE NOT EXISTS (
         SELECT 1
         FROM customer_identification ci
         WHERE ci.customer_identifier = cd.customer_identifier
     )
     """),

    ("Q40_completeness_score_per_customer",
     "Profile Completeness Score per Customer",
     """
     SELECT
         cd.customer_identifier,
         cd.customer_country,
         cd.customer_gender,
         cd.customer_status,
         (
             CASE WHEN cn.customer_identifier  IS NOT NULL THEN 1 ELSE 0 END +
             CASE WHEN ca.customer_identifier  IS NOT NULL THEN 1 ELSE 0 END +
             CASE WHEN cci.customer_identifier IS NOT NULL THEN 1 ELSE 0 END +
             CASE WHEN ci.customer_identifier  IS NOT NULL THEN 1 ELSE 0 END +
             CASE WHEN cp.customer_identifier  IS NOT NULL THEN 1 ELSE 0 END +
             CASE WHEN cct.customer_identifier IS NOT NULL THEN 1 ELSE 0 END
         ) AS completeness_score
     FROM customer_detail cd
     LEFT JOIN customer_name cn              ON cn.customer_identifier  = cd.customer_identifier
     LEFT JOIN customer_address ca           ON ca.customer_identifier  = cd.customer_identifier
     LEFT JOIN customer_contact_information cci ON cci.customer_identifier = cd.customer_identifier
     LEFT JOIN customer_identification ci   ON ci.customer_identifier  = cd.customer_identifier
     LEFT JOIN customer_proof_of_id cp      ON cp.customer_identifier  = cd.customer_identifier
     LEFT JOIN customer_classification_type cct ON cct.customer_identifier = cd.customer_identifier
     GROUP BY
         cd.customer_identifier, cd.customer_country,
         cd.customer_gender, cd.customer_status
     ORDER BY completeness_score DESC
     """),

    ("Q41_completeness_score_distribution",
     "Distribution of Profile Completeness Scores",
     """
     SELECT
         completeness_score,
         COUNT(*) AS customer_count
     FROM (
         SELECT
             cd.customer_identifier,
             (
                 CASE WHEN cn.customer_identifier  IS NOT NULL THEN 1 ELSE 0 END +
                 CASE WHEN ca.customer_identifier  IS NOT NULL THEN 1 ELSE 0 END +
                 CASE WHEN cci.customer_identifier IS NOT NULL THEN 1 ELSE 0 END +
                 CASE WHEN ci.customer_identifier  IS NOT NULL THEN 1 ELSE 0 END +
                 CASE WHEN cp.customer_identifier  IS NOT NULL THEN 1 ELSE 0 END +
                 CASE WHEN cct.customer_identifier IS NOT NULL THEN 1 ELSE 0 END
             ) AS completeness_score
         FROM customer_detail cd
         LEFT JOIN customer_name cn              ON cn.customer_identifier  = cd.customer_identifier
         LEFT JOIN customer_address ca           ON ca.customer_identifier  = cd.customer_identifier
         LEFT JOIN customer_contact_information cci ON cci.customer_identifier = cd.customer_identifier
         LEFT JOIN customer_identification ci   ON ci.customer_identifier  = cd.customer_identifier
         LEFT JOIN customer_proof_of_id cp      ON cp.customer_identifier  = cd.customer_identifier
         LEFT JOIN customer_classification_type cct ON cct.customer_identifier = cd.customer_identifier
         GROUP BY cd.customer_identifier
     ) AS scores
     GROUP BY completeness_score
     ORDER BY completeness_score DESC
     """),

    ("Q42_crud_flag_distribution",
     "CRUD Flag Distribution in Customer Detail",
     """
     SELECT
         crud_flag,
         COUNT(*) AS count
     FROM customer_detail
     GROUP BY crud_flag
     ORDER BY count DESC
     """),

    ("Q43_monthly_address_records",
     "Monthly Address Records Added (Last 3 Years)",
     """
     SELECT
         YEAR(audit_acceptance_time)  AS year,
         MONTH(audit_acceptance_time) AS month,
         COUNT(*)                     AS addresses_added
     FROM customer_address
     WHERE audit_acceptance_time >= DATE_SUB(NOW(), INTERVAL 3 YEAR)
     GROUP BY YEAR(audit_acceptance_time), MONTH(audit_acceptance_time)
     ORDER BY year DESC, month DESC
     """),

    ("Q44_contact_expiring_90_days",
     "Contact Records Expiring Within 90 Days",
     """
     SELECT
         customer_contact_type,
         COUNT(*) AS expiring_soon
     FROM customer_contact_information
     WHERE end_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 90 DAY)
     GROUP BY customer_contact_type
     ORDER BY expiring_soon DESC
     """),

    ("Q45_proof_of_id_expiring_90_days",
     "Proof of ID Records Expiring Within 90 Days",
     """
     SELECT
         customer_proof_of_id_type,
         COUNT(*) AS expiring_soon
     FROM customer_proof_of_id
     WHERE end_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 90 DAY)
     GROUP BY customer_proof_of_id_type
     ORDER BY expiring_soon DESC
     """),
]

def get_connection(host, port, user, password, database):
    log.info("Connecting to MySQL  host=%s  db=%s  user=%s", host, database, user)
    try:
        conn = mysql.connector.connect(
            host=host,
            port=int(port),
            user=user,
            password=password,
            database=database,
            connection_timeout=10,
            autocommit=True,
        )
        log.info("Connected successfully.")
        return conn
    except MySQLError as exc:
        log.error("Connection failed: %s", exc)
        sys.exit(1)



def run_queries(conn):
    """Execute every query and return a list of (qid, label, DataFrame)."""
    results = []
    total = len(QUERIES)
    for idx, (qid, label, sql) in enumerate(QUERIES, start=1):
        log.info("[%d/%d] Running: %s — %s", idx, total, qid, label)
        try:
            df = pd.read_sql(textwrap.dedent(sql).strip(), conn)
            log.info("  → %d rows returned", len(df))
        except Exception as exc:
            log.warning("  → FAILED (%s). Storing empty DataFrame.", exc)
            df = pd.DataFrame({"error": [str(exc)]})
        results.append((qid, label, df))
    return results

def export_csv(results, path):
    log.info("Writing CSV → %s", path)
    chunks = []
    for qid, label, df in results:
        header = pd.DataFrame([{col: f"=== {qid}: {label} ===" if i == 0 else ""
                                 for i, col in enumerate(df.columns)}])
        chunks.append(header)
        chunks.append(df)
        chunks.append(pd.DataFrame([{}]))          # blank separator row
    combined = pd.concat(chunks, ignore_index=True)
    combined.to_csv(path, index=False)
    log.info("CSV written (%d rows total).", len(combined))

def export_excel(results, path):
    log.info("Writing Excel → %s", path)

    # Colour palette
    HDR_FILL   = "0F2D4A"   # deep navy
    HDR_FONT   = "FFFFFF"
    ALT_FILL   = "EEF2F7"   # light blue-grey for alternating rows
    ACCENT     = "00B0FF"   # cyan accent for sheet tab

    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    thin = Side(style="thin", color="CFD8DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for qid, label, df in results:
        # Sheet name: max 31 chars, no special chars
        sheet_name = qid[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = ACCENT.replace("#", "")

        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=max(len(df.columns), 1))
        title_cell = ws.cell(row=1, column=1, value=f"{qid}  |  {label}")
        title_cell.font = Font(bold=True, size=12, color=HDR_FONT,
                               name="Calibri")
        title_cell.fill = PatternFill("solid", fgColor=HDR_FILL)
        title_cell.alignment = Alignment(horizontal="left",
                                         vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 22

        if df.empty:
            ws.cell(row=2, column=1, value="No data returned.")
            continue

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=str(col_name))
            cell.font      = Font(bold=True, color=HDR_FONT, name="Calibri",
                                  size=10)
            cell.fill      = PatternFill("solid", fgColor="1565C0")
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[2].height = 18

        for row_idx, row in enumerate(df.itertuples(index=False), start=3):
            fill_color = ALT_FILL if row_idx % 2 == 0 else "FFFFFF"
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill      = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")
                cell.border    = border
                cell.font      = Font(name="Calibri", size=9)

        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if not df.empty else 0
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws.freeze_panes = "A3"

    wb.save(path)
    log.info("Excel written (%d sheets).", len(results))


HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Customer Analytics Report</title>
<style>
  :root {{
    --navy:   #0F2D4A;
    --blue:   #1565C0;
    --cyan:   #00B0FF;
    --green:  #00C853;
    --light:  #F0F4F8;
    --border: #CFD8DC;
    --text:   #1A1A2E;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Segoe UI', Calibri, Arial, sans-serif;
    background: #f5f7fa;
    color: var(--text);
    font-size: 14px;
  }}

  /* ── Header ── */
  .report-header {{
    background: var(--navy);
    color: white;
    padding: 36px 48px 28px;
    border-bottom: 5px solid var(--cyan);
  }}
  .report-header h1 {{ font-size: 2rem; letter-spacing: 1px; margin-bottom: 6px; }}
  .report-header p  {{ color: #90CAF9; font-size: 0.9rem; }}

  /* ── Navigation sidebar ── */
  .layout {{ display: flex; min-height: 100vh; }}
  .sidebar {{
    width: 260px; min-width: 220px; flex-shrink: 0;
    background: var(--navy); padding: 24px 0;
    position: sticky; top: 0; height: 100vh; overflow-y: auto;
  }}
  .sidebar h2 {{
    color: var(--cyan); font-size: 0.75rem; letter-spacing: 2px;
    text-transform: uppercase; padding: 0 20px 10px;
  }}
  .sidebar a {{
    display: block; padding: 7px 20px; color: #B0C4DE;
    text-decoration: none; font-size: 0.78rem;
    border-left: 3px solid transparent;
    transition: all 0.15s;
  }}
  .sidebar a:hover {{
    color: white; background: rgba(255,255,255,0.07);
    border-left-color: var(--cyan);
  }}

  /* ── Main content ── */
  .main {{ flex: 1; padding: 32px 40px; max-width: 1200px; }}

  /* ── Query sections ── */
  .query-section {{
    background: white; border-radius: 8px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    margin-bottom: 32px; overflow: hidden;
  }}
  .query-header {{
    background: var(--navy); color: white;
    padding: 14px 20px; display: flex; align-items: center; gap: 12px;
  }}
  .query-id {{
    background: var(--cyan); color: var(--navy);
    font-size: 0.7rem; font-weight: 700; padding: 3px 8px;
    border-radius: 4px; white-space: nowrap; letter-spacing: 0.5px;
  }}
  .query-title {{ font-size: 0.95rem; font-weight: 600; }}
  .query-body {{ padding: 16px 20px; overflow-x: auto; }}
  .row-count {{
    font-size: 0.75rem; color: #78909C; margin-bottom: 10px;
  }}

  /* ── Tables ── */
  table {{
    border-collapse: collapse; width: 100%;
    font-size: 0.82rem; min-width: 300px;
  }}
  thead th {{
    background: var(--blue); color: white;
    padding: 9px 12px; text-align: center;
    font-weight: 600; white-space: nowrap;
    border: 1px solid #1255A8;
  }}
  tbody tr:nth-child(even) {{ background: var(--light); }}
  tbody tr:hover {{ background: #D9EAF7; }}
  tbody td {{
    padding: 7px 12px; border: 1px solid var(--border);
    text-align: center;
  }}
  tbody td:first-child {{ text-align: left; }}

  /* ── Empty state ── */
  .no-data {{
    color: #90A4AE; font-style: italic; padding: 12px 0;
  }}

  /* ── Footer ── */
  .report-footer {{
    text-align: center; padding: 24px;
    color: #90A4AE; font-size: 0.78rem;
    border-top: 1px solid var(--border); margin-top: 16px;
  }}

  /* ── Summary bar ── */
  .summary-bar {{
    display: flex; gap: 20px; flex-wrap: wrap;
    margin-bottom: 32px;
  }}
  .kpi-card {{
    background: white; border-radius: 8px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    padding: 18px 24px; flex: 1; min-width: 140px;
    border-top: 4px solid var(--cyan);
  }}
  .kpi-card .kpi-value {{ font-size: 1.8rem; font-weight: 700; color: var(--navy); }}
  .kpi-card .kpi-label {{ font-size: 0.78rem; color: #607D8B; margin-top: 4px; }}
</style>
</head>
<body>

<div class="report-header">
  <h1>Customer Analytics Report</h1>
  <p>Generated: {generated_at} &nbsp;|&nbsp; Database: {database} &nbsp;|&nbsp;
     {query_count} queries executed</p>
</div>

<div class="layout">

  <!-- Sidebar nav -->
  <nav class="sidebar">
    <h2>Queries</h2>
    {nav_links}
  </nav>

  <!-- Main -->
  <main class="main">

    <!-- Summary KPIs -->
    <div class="summary-bar">
      <div class="kpi-card">
        <div class="kpi-value">{query_count}</div>
        <div class="kpi-label">Queries Executed</div>
      </div>
      <div class="kpi-card">
        <div class="kpi-value">{success_count}</div>
        <div class="kpi-label">Successful</div>
      </div>
      <div class="kpi-card">
        <div class="kpi-value">{total_rows:,}</div>
        <div class="kpi-label">Total Rows Returned</div>
      </div>
    </div>

    <!-- Query results -->
    {sections}

  </main>
</div>

<div class="report-footer">
  Customer Analytics Report &nbsp;|&nbsp; {generated_at}
</div>
</body>
</html>
"""

def df_to_html_table(df):
    """Convert DataFrame to a styled HTML table string."""
    if df.empty:
        return '<p class="no-data">No data returned for this query.</p>'
    rows_html = ""
    for _, row in df.iterrows():
        cells = "".join(f"<td>{v}</td>" for v in row)
        rows_html += f"<tr>{cells}</tr>\n"
    headers = "".join(f"<th>{c}</th>" for c in df.columns)
    return (
        f"<table><thead><tr>{headers}</tr></thead>"
        f"<tbody>{rows_html}</tbody></table>"
    )

def export_html(results, path, database):
    log.info("Writing HTML → %s", path)
    generated_at  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    query_count   = len(results)
    success_count = sum(1 for _, _, df in results if "error" not in df.columns)
    total_rows    = sum(len(df) for _, _, df in results)

    nav_links = "\n".join(
        f'<a href="#{qid}">{qid}</a>'
        for qid, _, _ in results
    )

    sections_html = ""
    for qid, label, df in results:
        row_count_txt = f"{len(df):,} row(s) returned"
        table_html    = df_to_html_table(df)
        sections_html += f"""
        <div class="query-section" id="{qid}">
          <div class="query-header">
            <span class="query-id">{qid}</span>
            <span class="query-title">{label}</span>
          </div>
          <div class="query-body">
            <div class="row-count">{row_count_txt}</div>
            {table_html}
          </div>
        </div>
        """

    html = HTML_TEMPLATE.format(
        generated_at  = generated_at,
        database      = database,
        query_count   = query_count,
        success_count = success_count,
        total_rows    = total_rows,
        nav_links     = nav_links,
        sections      = sections_html,
    )
    Path(path).write_text(html, encoding="utf-8")
    log.info("HTML written.")


def parse_args():
    p = argparse.ArgumentParser(
        description="Run 45 customer analytics queries and export to CSV / Excel / HTML."
    )
    p.add_argument("--host",     default=os.getenv("DB_HOST",     "localhost"))
    p.add_argument("--port",     default=os.getenv("DB_PORT",     "3306"))
    p.add_argument("--user",     default=os.getenv("DB_USER",     "root"))
    p.add_argument("--password", default=os.getenv("DB_PASSWORD", ""))
    p.add_argument("--database", default=os.getenv("DB_NAME",     "lab4_db"))
    return p.parse_args()


def main():
    args    = parse_args()
    conn    = get_connection(
        args.host, args.port, args.user, args.password, args.database
    )
    results = run_queries(conn)
    conn.close()

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = REPORTS_DIR / f"customer_analytics_{ts}"

    export_csv(results,   str(base) + ".csv")
    export_excel(results, str(base) + ".xlsx")
    export_html(results,  str(base) + ".html", args.database)

    log.info("=" * 60)
    log.info("All done. Output files:")
    log.info("  CSV   → %s.csv",  base)
    log.info("  Excel → %s.xlsx", base)
    log.info("  HTML  → %s.html", base)
    log.info("=" * 60)


if __name__ == "__main__":
    main()
